import streamlit as st
import json
from pathlib import Path
from datetime import date
from typing import Dict, Any, List
import io
import csv
from collections import defaultdict

from openpyxl import Workbook
from openpyxl.utils import get_column_letter


APP_TITLE = "🎉 年中予算管理アプリ（家族イベント・冠婚葬祭）"
DATA_FILE = Path(__file__).parent / "hareday_data.json"
OLD_FILE = Path(__file__).parent / "otoshidama_data.json"


# =============================
# データ初期状態（最新版）
# =============================
def default_state() -> Dict[str, Any]:
    return {
        "members": ["家族", "子どもA"],
        "events": [
            # =========================
            # 季節・年中行事
            # =========================
            {"key": "otoshidama", "label": "お年玉", "group": "季節・年中行事"},
            {"key": "new_year", "label": "お正月", "group": "季節・年中行事"},
            {"key": "setsubun", "label": "節分", "group": "季節・年中行事"},
            {"key": "hinamatsuri", "label": "ひな祭り", "group": "季節・年中行事"},
            {"key": "white_day", "label": "ホワイトデー", "group": "季節・年中行事"},
            {"key": "graduation", "label": "卒園/卒業", "group": "季節・年中行事"},
            {"key": "entrance", "label": "入園/入学/進級", "group": "季節・年中行事"},
            {"key": "kodomo_no_hi", "label": "こどもの日", "group": "季節・年中行事"},
            {"key": "mothers_day", "label": "母の日", "group": "季節・年中行事"},
            {"key": "fathers_day", "label": "父の日", "group": "季節・年中行事"},
            {"key": "tanabata", "label": "七夕", "group": "季節・年中行事"},
            {"key": "summer_vacation", "label": "夏休み", "group": "季節・年中行事"},
            {"key": "obon", "label": "お盆", "group": "季節・年中行事"},
            {"key": "keiro_no_hi", "label": "敬老の日", "group": "季節・年中行事"},
            {"key": "halloween", "label": "ハロウィン", "group": "季節・年中行事"},
            {"key": "sports_day", "label": "運動会/発表会", "group": "季節・年中行事"},
            {"key": "shichigosan", "label": "七五三", "group": "季節・年中行事"},
            {"key": "christmas", "label": "クリスマス", "group": "季節・年中行事"},
            {"key": "birthday", "label": "誕生日", "group": "季節・年中行事"},
            {"key": "wedding_anniv", "label": "結婚記念日", "group": "季節・年中行事"},
            {"key": "year_end", "label": "年末（帰省/行事）", "group": "季節・年中行事"},

            # =========================
            # 冠婚葬祭
            # =========================
            {"key": "wedding_gift", "label": "結婚式（ご祝儀）", "group": "冠婚葬祭"},
            {"key": "funeral_condolence", "label": "葬儀（香典）", "group": "冠婚葬祭"},
            {"key": "memorial_service", "label": "法事/お墓参り", "group": "冠婚葬祭"},
            {"key": "hospital_visit", "label": "お見舞い", "group": "冠婚葬祭"},
            {"key": "birth_gift", "label": "出産祝い", "group": "冠婚葬祭"},
            {"key": "moving_gift", "label": "引っ越し/新築祝い", "group": "冠婚葬祭"},
            {"key": "return_gift", "label": "内祝い/お返し", "group": "冠婚葬祭"},

            # =========================
            # 旅行・大型支出
            # =========================
            {"key": "homecoming", "label": "帰省（交通費・手土産）", "group": "旅行・大型支出"},
            {"key": "family_trip", "label": "家族旅行", "group": "旅行・大型支出"},
            {"key": "leisure", "label": "レジャー", "group": "旅行・大型支出"},
            {"key": "special_shopping", "label": "特別な買い物（家具・家電）", "group": "旅行・大型支出"},
            {"key": "moving", "label": "引っ越し（費用全般）", "group": "旅行・大型支出"},

            # =========================
            # その他
            # =========================
            {"key": "other", "label": "その他", "group": "その他"},
        ],

        # 取引明細
        # {d, year, event_key, event_label, target, direction(in/out), category, amount, memo}
        "transactions": [],

        # 予算（年→イベント→金額）
        # 例: {"2025": {"birthday": 20000, "christmas": 30000}}
        "budgets": {},

        "migrated": False,
    }


# =============================
# ユーティリティ
# =============================
def safe_int(x, default=0) -> int:
    try:
        return int(x)
    except Exception:
        return default


def money(n: int) -> str:
    return f"¥{n:,}"


def load_json(path: Path) -> Dict[str, Any]:
    if path.exists():
        try:
            return json.loads(path.read_text(encoding="utf-8"))
        except Exception:
            pass
    return {}


def save_data(data: Dict[str, Any]):
    DATA_FILE.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")


def event_label_map(data: Dict[str, Any]) -> Dict[str, str]:
    return {e["key"]: e.get("label", e["key"]) for e in data.get("events", [])}


def event_group_map(data: Dict[str, Any]) -> Dict[str, str]:
    return {e["key"]: e.get("group", "その他") for e in data.get("events", [])}


def ensure_keys(data: Dict[str, Any]) -> Dict[str, Any]:
    """後方互換：足りないキーを補う"""
    base = default_state()
    for k, v in base.items():
        data.setdefault(k, v)
    if not data.get("members"):
        data["members"] = base["members"]
    if "budgets" not in data or data["budgets"] is None:
        data["budgets"] = {}
    if "transactions" not in data or data["transactions"] is None:
        data["transactions"] = []
    if "events" not in data or data["events"] is None:
        data["events"] = base["events"]
    return data


def sync_events_to_latest(data: Dict[str, Any]) -> Dict[str, Any]:
    """
    既存データの events を「最新版イベント定義」と同期する。
    - 既存のカスタムイベントは残す
    - 既存イベントに group が無いなら付与
    - 最新版にあるが既存に無いイベントは追加
    """
    latest = default_state()["events"]
    latest_by_key = {e["key"]: e for e in latest}

    current = data.get("events", []) or []
    current_by_key = {e.get("key"): e for e in current if e.get("key")}

    # 1) 既存イベントに group がなければ補完（最新版のgroupがあればそれ）
    for k, e in current_by_key.items():
        if "group" not in e or not e.get("group"):
            if k in latest_by_key and latest_by_key[k].get("group"):
                e["group"] = latest_by_key[k]["group"]
            else:
                e["group"] = "その他"
        if "label" not in e or not e.get("label"):
            e["label"] = k

    # 2) 最新版にあって既存に無いものを追加
    for k, e_latest in latest_by_key.items():
        if k not in current_by_key:
            current.append(dict(e_latest))

    # 3) 安定表示のため、標準イベント順（最新版順）→カスタム（それ以外）の順に並べる
    latest_order = [e["key"] for e in latest]
    custom = [e for e in current if e.get("key") not in set(latest_order)]
    reordered = []
    for k in latest_order:
        if k in current_by_key:
            reordered.append(current_by_key[k])
        else:
            # 追加されたはずだが、念のため
            reordered.append(dict(latest_by_key[k]))
    reordered.extend(custom)

    data["events"] = reordered
    return data


# =============================
# 旧お年玉データ移行
# =============================
def migrate_if_needed(raw: Dict[str, Any]) -> Dict[str, Any]:
    if not raw:
        return default_state()

    raw = ensure_keys(raw)

    # 旧形式チェック
    if all(k in raw for k in ["children", "received", "spent", "given"]):
        base = default_state()
        base["members"] = list(dict.fromkeys(["家族"] + raw.get("children", []))) or ["家族"]

        # received -> in
        for r in raw.get("received", []):
            d = r.get("d", str(date.today()))
            base["transactions"].append({
                "d": d,
                "year": safe_int(str(d)[:4], date.today().year),
                "event_key": "otoshidama",
                "event_label": "お年玉",
                "target": r.get("child", ""),
                "direction": "in",
                "category": f"もらった：{r.get('from','')}".strip() or "もらった",
                "amount": safe_int(r.get("amount", 0), 0),
                "memo": r.get("memo", ""),
            })

        # spent -> out
        for s in raw.get("spent", []):
            d = s.get("d", str(date.today()))
            base["transactions"].append({
                "d": d,
                "year": safe_int(str(d)[:4], date.today().year),
                "event_key": "otoshidama",
                "event_label": "お年玉",
                "target": s.get("child", ""),
                "direction": "out",
                "category": s.get("category", "支出"),
                "amount": safe_int(s.get("amount", 0), 0),
                "memo": s.get("memo", ""),
            })

        # given -> out (target=相手)
        for g in raw.get("given", []):
            d = g.get("d", str(date.today()))
            base["transactions"].append({
                "d": d,
                "year": safe_int(str(d)[:4], date.today().year),
                "event_key": "otoshidama",
                "event_label": "お年玉",
                "target": g.get("to", "相手"),
                "direction": "out",
                "category": "渡したお年玉",
                "amount": safe_int(g.get("amount", 0), 0),
                "memo": g.get("memo", ""),
            })

        base["migrated"] = True
        return base

    return raw


# =============================
# 予算ヘルパー
# =============================
def get_budget(data: Dict[str, Any], year: int, event_key: str) -> int:
    return int(data.get("budgets", {}).get(str(year), {}).get(event_key, 0) or 0)


def set_budget(data: Dict[str, Any], year: int, event_key: str, amount: int) -> None:
    data.setdefault("budgets", {})
    data["budgets"].setdefault(str(year), {})
    data["budgets"][str(year)][event_key] = int(amount)


# =============================
# 集計
# =============================
def sum_filter(data: Dict[str, Any], year: int = None, event_key: str = None, direction: str = None) -> int:
    total = 0
    for t in data.get("transactions", []):
        if year is not None and t.get("year") != year:
            continue
        if event_key is not None and event_key != "" and t.get("event_key") != event_key:
            continue
        if direction is not None and direction != "" and t.get("direction") != direction:
            continue
        total += safe_int(t.get("amount", 0), 0)
    return total


def sum_event_out(data: Dict[str, Any], year: int, event_key: str) -> int:
    return sum_filter(data, year=year, event_key=event_key, direction="out")


def available_years(data: Dict[str, Any]) -> List[int]:
    ys = sorted({safe_int(t.get("year", 0), 0) for t in data.get("transactions", []) if safe_int(t.get("year", 0), 0) > 0})
    if not ys:
        ys = [date.today().year]
    return ys


# =============================
# エクスポート
# =============================
def export_csv(transactions: List[Dict[str, Any]]) -> bytes:
    output = io.StringIO()
    writer = csv.DictWriter(
        output,
        fieldnames=["d", "year", "event_label", "event_key", "target", "direction", "category", "amount", "memo"]
    )
    writer.writeheader()
    for t in transactions:
        writer.writerow({
            "d": t.get("d", ""),
            "year": t.get("year", ""),
            "event_label": t.get("event_label", ""),
            "event_key": t.get("event_key", ""),
            "target": t.get("target", ""),
            "direction": t.get("direction", ""),
            "category": t.get("category", ""),
            "amount": t.get("amount", 0),
            "memo": t.get("memo", ""),
        })
    return output.getvalue().encode("utf-8-sig")


def autosize(ws):
    for col in ws.columns:
        length = max(len(str(c.value)) if c.value is not None else 0 for c in col)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(48, max(10, length + 2))


def export_xlsx(data: Dict[str, Any]) -> bytes:
    wb = Workbook()

    # Transactions
    ws = wb.active
    ws.title = "Transactions"
    ws.append(["日付", "年", "イベント", "event_key", "対象", "収支", "内容", "金額", "メモ"])
    for t in data.get("transactions", []):
        ws.append([
            t.get("d", ""),
            t.get("year", ""),
            t.get("event_label", ""),
            t.get("event_key", ""),
            t.get("target", ""),
            t.get("direction", ""),
            t.get("category", ""),
            t.get("amount", 0),
            t.get("memo", ""),
        ])
    autosize(ws)

    # Budgets
    ws_b = wb.create_sheet("Budgets")
    ws_b.append(["年", "event_key", "イベント", "グループ", "予算(円)"])
    ev_map = event_label_map(data)
    gr_map = event_group_map(data)
    for y, m in (data.get("budgets", {}) or {}).items():
        for ek, amt in (m or {}).items():
            ws_b.append([y, ek, ev_map.get(ek, ek), gr_map.get(ek, "その他"), amt])
    autosize(ws_b)

    # Summary
    ws2 = wb.create_sheet("Summary")
    ws2.append(["年", "グループ", "イベント", "予算", "支出", "残り", "収入", "差額(in-out)"])
    years = available_years(data)
    for y in years:
        for e in data.get("events", []):
            ek = e["key"]
            label = e.get("label", ek)
            group = e.get("group", "その他")
            budget = get_budget(data, y, ek)
            out_total = sum_filter(data, y, ek, "out")
            in_total = sum_filter(data, y, ek, "in")
            if budget == 0 and out_total == 0 and in_total == 0:
                continue
            ws2.append([y, group, label, budget, out_total, budget - out_total, in_total, in_total - out_total])
    autosize(ws2)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# =============================
# UI: 広告（A8）
# =============================
def render_a8_corner_ad():
    """
    右下固定のPR枠。A8の「画像バナー（a+img）」が安定。
    script系しか無い場合は動かないことがあるので、まず画像バナー推奨。
    """
    # ▼▼▼ ここにA8の広告HTML（画像バナー推奨）を貼ってね ▼▼▼
    A8_HTML = ""  # 例：<a ...><img .../></a>

    if not A8_HTML.strip():
        A8_HTML = '<div style="opacity:0.8;">広告枠（A8タグをここに貼る）</div>'

    st.markdown(
        f"""
<style>
.pr-box {{
  position: fixed;
  right: 12px;
  bottom: 12px;
  width: 240px;
  padding: 10px 12px;
  border: 1px solid rgba(255,255,255,0.15);
  border-radius: 12px;
  background: rgba(20,20,20,0.60);
  backdrop-filter: blur(6px);
  z-index: 9999;
  font-size: 12px;
}}
.pr-label {{
  font-size: 11px;
  opacity: 0.7;
  margin-bottom: 6px;
}}
.pr-content img {{
  max-width: 100%;
  height: auto;
  border-radius: 8px;
}}
</style>
<div class="pr-box">
  <div class="pr-label">PR</div>
  <div class="pr-content">{A8_HTML}</div>
</div>
""",
        unsafe_allow_html=True
    )


# =============================
# アプリ開始
# =============================
st.set_page_config(page_title=APP_TITLE, layout="wide")
st.title(APP_TITLE)
st.caption("年中行事・冠婚葬祭・特別支出を“年単位の予算”で管理。CSV/Excelで持ち出しOK。")

raw = load_json(DATA_FILE) or load_json(OLD_FILE)
data = migrate_if_needed(raw)
data = ensure_keys(data)

# ★ここがポイント：イベント定義を最新版へ同期（group補完・不足追加・並び安定）
data = sync_events_to_latest(data)

# 保存（同期内容をJSONに反映）
save_data(data)

ev_map = event_label_map(data)
gr_map = event_group_map(data)

# 右下広告（A8）
render_a8_corner_ad()

# -----------------------------
# サイドバー：メンバー/イベント管理
# -----------------------------
with st.sidebar:
    st.subheader("👪 メンバー")

    # 追加
    new_member = st.text_input("追加", placeholder="例：子どもB / 妻 / 自分")
    if st.button("メンバー追加", use_container_width=True):
        name = (new_member or "").strip()
        if name and name not in data["members"]:
            data["members"].append(name)
            save_data(data)
            st.rerun()

    # 削除（取引が無ければ削除OK）
    if data["members"]:
        del_target = st.selectbox("削除するメンバー", options=data["members"])
        used = any(t.get("target") == del_target for t in data.get("transactions", []))
        if used:
            st.warning("このメンバーは取引履歴があるため削除できません。")
        else:
            if st.button("このメンバーを削除", type="secondary", use_container_width=True):
                data["members"].remove(del_target)
                save_data(data)
                st.success("削除しました")
                st.rerun()

    st.divider()
    st.subheader("🎉 イベント（追加）")

    new_event = st.text_input("イベント名", placeholder="例：受験 / 推し活 / 写真撮影")
    new_group = st.selectbox("グループ", ["季節・年中行事", "冠婚葬祭", "旅行・大型支出", "その他"], index=3)
    if st.button("イベント追加", use_container_width=True):
        label = (new_event or "").strip()
        if label:
            key = label.lower().replace(" ", "_")
            existed = {e["key"] for e in data["events"]}
            if key in existed:
                key = f"{key}_{len(existed)+1}"
            data["events"].append({"key": key, "label": label, "group": new_group})
            save_data(data)
            st.rerun()

    st.divider()
    st.subheader("💾")
    st.write(f"保存先：`{DATA_FILE.name}`")
    if st.button("データ初期化（全削除）", type="secondary", use_container_width=True):
        data = default_state()
        save_data(data)
        st.rerun()


tab1, tab2, tab3 = st.tabs(["① 記録する", "② 予算と集計", "③ エクスポート"])


# =============================
# ① 記録する
# =============================
with tab1:
    st.subheader("📝 入出金を追加（10秒設計）")

    d = st.date_input("日付", value=date.today())
    year = d.year

    event_key = st.selectbox(
        "イベント",
        options=[e["key"] for e in data["events"]],
        format_func=lambda k: ev_map.get(k, k)
    )
    target = st.selectbox("対象", options=data["members"])
    direction_ui = st.radio("種類", ["支出（out）", "収入（in）"], horizontal=True)
    direction = "out" if "out" in direction_ui else "in"

    category = st.text_input("内容（カテゴリ）", placeholder="例：プレゼント / ご祝儀 / 香典 / 外食 / ケーキ")
    amount = st.number_input("金額（円）", min_value=0, step=500, value=1000)
    memo = st.text_input("メモ（任意）", placeholder="例：店名 / 誰から / 来年同額")

    c1, c2 = st.columns([1, 1])
    with c1:
        if st.button("＋ 追加", use_container_width=True):
            if (category or "").strip() and int(amount) > 0:
                data["transactions"].append({
                    "d": str(d),
                    "year": year,
                    "event_key": event_key,
                    "event_label": ev_map.get(event_key, event_key),
                    "target": target,
                    "direction": direction,
                    "category": category.strip(),
                    "amount": int(amount),
                    "memo": (memo or "").strip(),
                })
                save_data(data)
                st.success("追加したよ！")
                st.rerun()
            else:
                st.warning("「内容」と「金額」を入れてね。")

    with c2:
        if st.button("↩︎ 入力をクリア（見た目だけ）", use_container_width=True):
            st.rerun()

    st.divider()
    st.subheader("📚 直近の明細（最新10件）")
    rows = data.get("transactions", [])[-10:][::-1]
    if not rows:
        st.info("まだ記録がないよ。")
    else:
        for t in rows:
            with st.container(border=True):
                sign = "＋" if t.get("direction") == "in" else "−"
                st.write(
                    f"**{t.get('d','')}**  "
                    f"{t.get('event_label','')} / {t.get('target','')} / {t.get('category','')}  "
                    f"—  **{sign}{money(safe_int(t.get('amount',0),0))}**"
                )
                if t.get("memo"):
                    st.caption(t["memo"])


# =============================
# ② 予算と集計（グループ別表示）
# =============================
with tab2:
    st.subheader("📊 年別サマリー")
    years = available_years(data)
    sel_year = st.selectbox("年", options=years, index=len(years) - 1)

    total_in = sum_filter(data, year=sel_year, event_key="", direction="in")
    total_out = sum_filter(data, year=sel_year, event_key="", direction="out")

    a, b, c = st.columns(3)
    a.metric("収入 合計", money(total_in))
    b.metric("支出 合計", money(total_out))
    c.metric("差額（in-out）", money(total_in - total_out))

    st.divider()

    # 表示オプション
    opt1, opt2, opt3 = st.columns([1, 1, 1])
    with opt1:
        show_only_used = st.checkbox("使ったイベントだけ表示", value=False)
    with opt2:
        expand_all = st.checkbox("全部ひらく（重いので注意）", value=False)
    with opt3:
        show_income_in_event = st.checkbox("イベント内で収入も表示", value=False)

    st.subheader("🎯 イベント別：予算 / 支出 / 残り（グループ別）")
    st.caption("まずは“予算だけ”入れていけばOK。明細は後からでも回る。")

    # グルーピング（順序固定）
    group_order = ["季節・年中行事", "冠婚葬祭", "旅行・大型支出", "その他"]
    grouped = defaultdict(list)
    for e in data.get("events", []):
        grouped[e.get("group", "その他")].append(e)

    def should_show_event(ek: str) -> bool:
        if not show_only_used:
            return True
        out_total = sum_event_out(data, sel_year, ek)
        in_total = sum_filter(data, year=sel_year, event_key=ek, direction="in")
        budget = get_budget(data, sel_year, ek)
        return (out_total != 0) or (in_total != 0) or (budget != 0)

    # グループごとの合計を出す
    for gname in group_order:
        events = grouped.get(gname, [])
        if not events:
            continue

        # グループ合計
        g_budget = 0
        g_out = 0
        g_in = 0
        for e in events:
            ek = e["key"]
            if not should_show_event(ek):
                continue
            g_budget += get_budget(data, sel_year, ek)
            g_out += sum_event_out(data, sel_year, ek)
            g_in += sum_filter(data, year=sel_year, event_key=ek, direction="in")

        # 「使ったイベントだけ表示」かつ、全部0ならグループ自体を出さない
        if show_only_used and (g_budget == 0 and g_out == 0 and g_in == 0):
            continue

        st.subheader(f"📂 {gname}")
        colx, coly, colz = st.columns(3)
        colx.metric("グループ予算", money(g_budget))
        coly.metric("グループ支出", money(g_out))
        colz.metric("グループ残り", money(g_budget - g_out))

        for e in events:
            ek = e["key"]
            label = e.get("label", ek)

            if not should_show_event(ek):
                continue

            out_total = sum_event_out(data, sel_year, ek)
            in_total = sum_filter(data, year=sel_year, event_key=ek, direction="in")
            budget = get_budget(data, sel_year, ek)
            remaining = budget - out_total

            header = f"{label}  |  予算 {money(budget)} / 支出 {money(out_total)} / 残り {money(remaining)}"
            with st.expander(header, expanded=expand_all):
                col1, col2, col3 = st.columns([1.2, 1, 1])

                with col1:
                    new_budget = st.number_input(
                        "今年の予算（円）",
                        min_value=0,
                        step=1000,
                        value=int(budget),
                        key=f"budget_{sel_year}_{ek}"
                    )
                    if st.button("予算を保存", key=f"save_budget_{sel_year}_{ek}", use_container_width=True):
                        set_budget(data, sel_year, ek, int(new_budget))
                        save_data(data)
                        st.success("保存したよ！")
                        st.rerun()

                with col2:
                    st.metric("支出合計", money(out_total))
                    st.metric("残り予算", money(remaining))

                with col3:
                    if show_income_in_event and in_total:
                        st.metric("収入合計", money(in_total))
                    if budget > 0:
                        used_ratio = min(1.0, out_total / budget) if budget else 0.0
                        st.progress(used_ratio)
                        used_pct = int((out_total / budget) * 100) if budget else 0
                        st.caption(f"消化率：{used_pct}%")

        st.divider()

    # -------------------------
    # 明細フィルタ（この年）
    # -------------------------
    st.subheader("🔎 明細フィルタ（この年）")

    # グループで絞り込み
    groups_for_filter = ["（全部）"] + group_order
    g_choice = st.selectbox("グループで絞り込み", options=groups_for_filter)

    # イベントで絞り込み
    ev_keys = [e["key"] for e in data["events"]]
    ev_choice = st.selectbox(
        "イベントで絞り込み",
        options=["（全部）"] + ev_keys,
        format_func=lambda k: "（全部）" if k == "（全部）" else ev_map.get(k, k)
    )

    target_choice = st.selectbox("対象で絞り込み", options=["（全部）"] + data["members"])
    dir_choice = st.selectbox("収支で絞り込み", options=["（全部）", "in", "out"])

    def match(t):
        if t.get("year") != sel_year:
            return False

        # グループ絞り込み
        if g_choice != "（全部）":
            ek = t.get("event_key")
            if gr_map.get(ek, "その他") != g_choice:
                return False

        if ev_choice != "（全部）" and t.get("event_key") != ev_choice:
            return False
        if target_choice != "（全部）" and t.get("target") != target_choice:
            return False
        if dir_choice != "（全部）" and t.get("direction") != dir_choice:
            return False
        return True

    filtered = [t for t in data.get("transactions", []) if match(t)]
    if not filtered:
        st.info("条件に合う明細がないよ。")
    else:
        # 新しい順に表示
        for idx, t in enumerate(filtered[::-1], start=1):
            with st.container(border=True):
                sign = "＋" if t.get("direction") == "in" else "−"
                st.write(
                    f"**{t.get('d','')}**  "
                    f"{t.get('event_label','')} / {t.get('target','')} / {t.get('category','')}  "
                    f"—  **{sign}{money(safe_int(t.get('amount',0),0))}**"
                )
                if t.get("memo"):
                    st.caption(t["memo"])

                if st.button("削除（この行）", key=f"del_tx_{sel_year}_{idx}"):
                    # 同一dictを見つけて削除
                    try:
                        real_idx = data["transactions"].index(t)
                        data["transactions"].pop(real_idx)
                        save_data(data)
                        st.rerun()
                    except ValueError:
                        # 念のため保険
                        st.warning("削除に失敗。再読み込み後にもう一度試してね。")


# =============================
# ③ エクスポート
# =============================
with tab3:
    st.subheader("📦 エクスポート（CSV / Excel）")

    # event_label を最新に整える（イベント名変更時の保険）
    ev_map_now = event_label_map(data)
    for t in data.get("transactions", []):
        k = t.get("event_key")
        if k:
            t["event_label"] = ev_map_now.get(k, t.get("event_label", k))

    csv_bytes = export_csv(data.get("transactions", []))
    xlsx_bytes = export_xlsx(data)

    st.download_button(
        "⬇️ CSVをダウンロード（Excelで開ける）",
        data=csv_bytes,
        file_name="hareday_transactions.csv",
        mime="text/csv",
        use_container_width=True,
    )

    st.download_button(
        "⬇️ Excel（XLSX）をダウンロード（Transactions / Budgets / Summary）",
        data=xlsx_bytes,
        file_name="hareday_export.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True,
    )

    st.divider()
    st.markdown("### データ（JSON）")
    st.code(json.dumps(data, ensure_ascii=False, indent=2), language="json")
